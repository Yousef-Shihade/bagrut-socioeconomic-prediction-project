"""
io_load.py — Step 6 input loading (budget xlsx + Step-4 table) & robust parsing.

Project: Predicting Israeli High School Bagrut Success Using Socioeconomic Data
Authors: Yousef Shehade & Shada Esawi

The Ministry budget workbook ships a malformed styles.xml (non-aRGB theme
colours) that makes a vanilla `openpyxl.load_workbook` raise
``ValueError: Colors must be aRGB hex values``. We monkeypatch openpyxl's RGB
descriptor with a lenient validator *before* opening the file, so style parsing
can never abort the load. We then read in ``read_only`` + ``data_only`` mode
(values, not formulae) and hand back a clean DataFrame.
"""
from __future__ import annotations

import re
import warnings
from pathlib import Path
from typing import Any

import pandas as pd
import yaml

ROOT = Path(__file__).resolve().parents[1]
CONFIG_PATH = ROOT / "config.yaml"


# --------------------------------------------------------------------------- #
# openpyxl hardening                                                          #
# --------------------------------------------------------------------------- #
def _patch_openpyxl_colors() -> None:
    """Make openpyxl tolerate the budget file's invalid theme colours."""
    import openpyxl.styles.colors as colors

    if getattr(colors.RGB, "_step6_patched", False):
        return
    _orig = colors.RGB.__set__

    def _lenient(self, instance, value):  # noqa: ANN001
        try:
            _orig(self, instance, value)
        except (ValueError, TypeError):
            _orig(self, instance, "00000000")

    colors.RGB.__set__ = _lenient
    colors.RGB._step6_patched = True


def load_config(path: Path | str = CONFIG_PATH) -> dict[str, Any]:
    with open(path, "r", encoding="utf-8") as fh:
        return yaml.safe_load(fh)


def resolve(rel_path: str) -> Path:
    return (ROOT / rel_path).resolve()


# --------------------------------------------------------------------------- #
# loaders                                                                     #
# --------------------------------------------------------------------------- #
def load_budget_raw(cfg: dict[str, Any]) -> pd.DataFrame:
    """Read the budget workbook into a DataFrame, header on row 1.

    The grand-totals row (first data row, label ``סה"כ``) is dropped. Hebrew
    column headers are whitespace-normalised (the source has stray double
    spaces, e.g. ``'עלות  שעות הוראה'``) so the config column map matches.
    """
    _patch_openpyxl_colors()
    import openpyxl

    path = resolve(cfg["paths"]["budget_xlsx"])
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[cfg["io"]["budget_sheet"]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

    header = [re.sub(r"\s+", " ", str(c)).strip() if c is not None else c
              for c in rows[0]]
    df = pd.DataFrame(rows[1:], columns=header)

    # Drop the grand-totals row (label in the first/semel-year columns).
    totals = cfg["budget_columns"]["totals_label"]
    mask_total = df.apply(lambda r: r.astype(str).str.strip().eq(totals).any(), axis=1)
    df = df[~mask_total].reset_index(drop=True)
    return df


def load_cleaned(cfg: dict[str, Any]) -> pd.DataFrame:
    """Load the finalized Step-4 modeling table (the 2-dataset matrix)."""
    df = pd.read_csv(resolve(cfg["paths"]["cleaned_in"]),
                     encoding=cfg["io"]["encoding"])
    flag = cfg["features"].get("exclude_outlier_flag")
    if flag and flag in df.columns:
        df = df[~df[flag].astype(bool)].reset_index(drop=True)
    return df


def load_step5_leaderboard(cfg: dict[str, Any]) -> pd.DataFrame:
    """Step-5 untuned tournament leaderboard (context for the comparison)."""
    return pd.read_csv(resolve(cfg["paths"]["step5_leaderboard"]),
                       encoding="utf-8-sig")


if __name__ == "__main__":
    cfg = load_config()
    b = load_budget_raw(cfg)
    s4 = load_cleaned(cfg)
    print(f"[io_load] budget raw: {b.shape}; step4 cleaned (no outliers): {s4.shape}")
    print("[io_load] budget cols present:",
          [v for v in cfg["budget_columns"].values() if v in b.columns])
